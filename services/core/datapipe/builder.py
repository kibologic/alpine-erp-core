"""
datapipe.builder — Excel Template Builder
==========================================
Generates XLSX templates from a SchemaDefinitionObject + manifest config.

Output format (per sheet):
  Row 1: Embedded metadata (schema_version, timestamp) — hidden-ish marker
  Row 2: Column headers (from live schema, NOT hardcoded)
  Row 3: Type hints     (e.g. "Type: text | max 255", "Type: decimal")
  Row 4: Required/Optional markers
  Row 5+: Data rows (if exporting data) OR sample row (if template only)

Locked/excluded columns are styled differently and skipped from user input.
FK columns show their human-readable display label (from manifest).
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .introspect import ColumnDefinition, ColumnKind, SchemaDefinitionObject
from .mapper import FKResolution


# ─── Style Constants ──────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1E293B")   # dark slate
_HEADER_FONT  = Font(bold=True, color="F8FAFC")
_LOCKED_FILL  = PatternFill("solid", fgColor="374151")   # grey
_LOCKED_FONT  = Font(color="9CA3AF", italic=True)
_HINT_FONT    = Font(color="6B7280", italic=True, size=9)
_REQUIRED_FONT= Font(color="EF4444", bold=True, size=9)
_OPTIONAL_FONT= Font(color="9CA3AF", size=9)
_SAMPLE_FILL  = PatternFill("solid", fgColor="F0FDF4")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _kind_hint(col: ColumnDefinition, fk_res: Optional[FKResolution]) -> str:
    """Human-readable type hint string for Row 3."""
    if fk_res:
        return f"FK → {fk_res.model.__tablename__}.{fk_res.match_field}"
    if col.kind == ColumnKind.ENUM:
        opts = " | ".join(col.enum_values[:6])
        return f"Enum: {opts}"
    if col.kind == ColumnKind.TEXT and col.max_length:
        return f"Text (max {col.max_length})"
    return {
        ColumnKind.TEXT:    "Text",
        ColumnKind.INTEGER: "Integer",
        ColumnKind.DECIMAL: "Decimal (e.g. 10.50)",
        ColumnKind.BOOLEAN: "true / false",
        ColumnKind.DATE:    "Date (YYYY-MM-DD)",
        ColumnKind.DATETIME:"Datetime (YYYY-MM-DD HH:MM:SS)",
        ColumnKind.UUID:    "UUID",
        ColumnKind.JSON:    "JSON (advanced)",
    }.get(col.kind, col.kind.value)


def _sample_value(col: ColumnDefinition, fk_res: Optional[FKResolution]) -> Any:
    """Conservative sample value for the sample row."""
    if fk_res:
        return f"<{fk_res.display_label}>"
    return {
        ColumnKind.TEXT:    "Example text",
        ColumnKind.INTEGER: 1,
        ColumnKind.DECIMAL: 9.99,
        ColumnKind.BOOLEAN: True,
        ColumnKind.DATE:    "2024-01-01",
        ColumnKind.DATETIME:"2024-01-01 00:00:00",
        ColumnKind.ENUM:    col.enum_values[0] if col.enum_values else "value",
    }.get(col.kind, "")


# ─── Column Filtering ─────────────────────────────────────────────────────────

def _build_visible_columns(
    sdo: SchemaDefinitionObject,
    locked_fields: set[str],
    excluded_fields: set[str],
    fk_resolutions: dict[str, FKResolution],
    show_locked: bool = False,
) -> list[tuple[ColumnDefinition, bool, Optional[FKResolution]]]:
    """
    Returns list of (col_def, is_locked, fk_resolution_or_None) for visible columns.
    Skips excluded fields and (by default) locked fields entirely.
    Set show_locked=True to include locked columns greyed out (for exports).
    """
    result = []
    for col in sdo.columns:
        if col.name in excluded_fields:
            continue
        is_locked = col.name in locked_fields or col.primary_key
        # Skip locked columns from import templates (users can't and shouldn't set them)
        if is_locked and not show_locked:
            continue
        fk_res = fk_resolutions.get(col.name)
        result.append((col, is_locked, fk_res))
    return result


def _header_label(col: ColumnDefinition, fk_res: Optional[FKResolution]) -> str:
    """Column header label — use FK display label if available."""
    if fk_res:
        return fk_res.display_label
    # Humanize: snake_case → Title Case (strip _id suffix for FK cols)
    name = col.name
    if name.endswith("_id"):
        name = name[:-3]
    return name.replace("_", " ").title()


# ─── Public API ───────────────────────────────────────────────────────────────

def build_template(
    sdo: SchemaDefinitionObject,
    locked_fields: set[str],
    excluded_fields: set[str],
    fk_resolutions: dict[str, FKResolution],
    data_rows: Optional[list[dict[str, Any]]] = None,
    schema_version: str = "1",
    include_sample: bool = True,
    show_locked: bool = False,
) -> bytes:
    """
    Generate an XLSX template from a SchemaDefinitionObject.

    Args:
        sdo:              Schema definition for the table.
        locked_fields:    Fields that are system-managed (displayed but protected).
        excluded_fields:  Fields completely hidden from the template.
        fk_resolutions:   { fk_column_name: FKResolution } for human-readable FK display.
        data_rows:        If provided, export actual data instead of a sample row.
        schema_version:   Embedded in metadata for round-trip drift detection.
        include_sample:   Add a sample row (ignored if data_rows provided).

    Returns:
        XLSX file as bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sdo.table_name.replace("_", " ").title()

    visible = _build_visible_columns(sdo, locked_fields, excluded_fields, fk_resolutions, show_locked=show_locked)

    # ── Row 1: Embedded Metadata ──────────────────────────────────────────────
    meta_values = [
        "__meta__",
        f"schema_version={sdo.table_name}_v{schema_version}",
        f"export_timestamp={datetime.utcnow().isoformat()}",
        f"table={sdo.table_name}",
    ]
    ws.append(meta_values)
    for cell in ws[1]:
        cell.font = Font(color="CBD5E1", size=8, italic=True)

    # ── Row 2: Headers ────────────────────────────────────────────────────────
    headers = [_header_label(col, fk_res) for col, _, fk_res in visible]
    ws.append(headers)
    header_row = ws[2]
    for col_idx, (cell, (col_def, is_locked, _)) in enumerate(zip(header_row, visible), start=1):
        if is_locked:
            cell.fill = _LOCKED_FILL
            cell.font = _LOCKED_FONT
        else:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(headers[col_idx - 1]) + 4)

    # ── Row 3: Type Hints ─────────────────────────────────────────────────────
    hints = [_kind_hint(col, fk_res) for col, _, fk_res in visible]
    ws.append(hints)
    for cell in ws[3]:
        cell.font = _HINT_FONT

    # ── Row 4: Required / Optional ────────────────────────────────────────────
    required_markers = []
    for col_def, is_locked, _ in visible:
        if is_locked:
            required_markers.append("system")
        elif not col_def.nullable and col_def.default is None and not col_def.primary_key:
            required_markers.append("✓ required")
        else:
            required_markers.append("○ optional")
    ws.append(required_markers)
    for cell, marker in zip(ws[4], required_markers):
        if "required" in marker:
            cell.font = _REQUIRED_FONT
        else:
            cell.font = _OPTIONAL_FONT

    # ── Data Rows / Sample ────────────────────────────────────────────────────
    if data_rows:
        for data_row in data_rows:
            row_vals = []
            for col_def, is_locked, fk_res in visible:
                # For FK columns with resolutions, show the human-readable value
                # that was stored via the resolution (must be pre-resolved to display)
                val = data_row.get(col_def.name)
                if val is None:
                    row_vals.append("")
                elif isinstance(val, bool):
                    row_vals.append(str(val).lower())
                else:
                    row_vals.append(val)
            ws.append(row_vals)
    elif include_sample:
        sample = [_sample_value(col, fk_res) for col, _, fk_res in visible]
        ws.append(sample)
        for cell in ws[5]:
            cell.fill = _SAMPLE_FILL
            cell.font = Font(italic=True, color="4B5563")

    # ── Freeze panes below header + type rows ─────────────────────────────────
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
