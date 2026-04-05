"""
datapipe.parser — Excel Import Parser
======================================
Reads an uploaded XLSX file and yields normalized RawRow dicts.

Responsibilities:
  - Read sheets by name
  - Extract and normalize headers (strip, lowercase)
  - Ignore unknown columns gracefully
  - Detect which required columns are missing
  - Yield row dicts keyed by normalized header

No type coercion here. That's coerce.py's job.
No ERP knowledge.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl


# ─── Types ────────────────────────────────────────────────────────────────────

@dataclass
class ParsedSheet:
    name: str
    headers: list[str]          # normalized (stripped, lowercased)
    raw_headers: list[str]      # original as found in file
    rows: list[dict[str, Any]]  # { normalized_header: value }
    meta: dict[str, str]        # embedded metadata from the file (row 1 style)


@dataclass
class ParseResult:
    sheets: dict[str, ParsedSheet]   # sheet_name → ParsedSheet
    meta: dict[str, str]             # file-level metadata (schema_version, etc.)
    warnings: list[str]


# ─── Metadata Row ─────────────────────────────────────────────────────────────

_META_ROW_MARKER = "__meta__"
_META_FIRST_COL  = "alpine_meta"


def _extract_meta_row(ws) -> dict[str, str]:
    """
    Look for an optional metadata row anywhere in the first 5 rows.
    Format: first cell = "__meta__", rest are key=value pairs.
    Example: __meta__ | schema_version=products_v2 | export_timestamp=2024-01-01
    """
    meta: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row and str(row[0] or "").strip().lower() == _META_ROW_MARKER:
            for cell in row[1:]:
                if cell and "=" in str(cell):
                    k, _, v = str(cell).partition("=")
                    meta[k.strip()] = v.strip()
            break
    return meta


# ─── Header Detection ─────────────────────────────────────────────────────────

def _find_header_row(ws, max_scan: int = 6) -> Optional[int]:
    """
    Find the first non-meta row that looks like a header.
    Returns the 1-based row number, or None if not found.
    A header row is detected as a row where the first cell is a non-empty string
    that doesn't start with '__'.
    """
    for row_idx in range(1, max_scan + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        if not any(row):
            continue
        first = str(row[0] or "").strip()
        if first and not first.startswith("__") and not first.lower().startswith("type:") and not first.lower().startswith("required"):
            return row_idx
    return None


# ─── Public API ───────────────────────────────────────────────────────────────

def parse_workbook(
    content: bytes,
    sheet_name: Optional[str] = None,
    known_fields: Optional[set[str]] = None,
) -> ParseResult:
    """
    Parse an XLSX workbook and return structured ParseResult.

    Args:
        content:      Raw bytes of the XLSX file.
        sheet_name:   If provided, only parse this sheet. Otherwise parse all.
        known_fields: If provided, emit warnings for unknown columns.

    Returns:
        ParseResult containing all parsed sheets and file-level metadata.
    """
    wb       = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    warnings: list[str] = []
    sheets:   dict[str, ParsedSheet] = {}
    file_meta: dict[str, str] = {}

    target_sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sname in target_sheets:
        if sname not in wb.sheetnames:
            warnings.append(f"Sheet '{sname}' not found in workbook")
            continue

        ws = wb[sname]

        # Extract any embedded metadata
        sheet_meta = _extract_meta_row(ws)
        if not file_meta:
            file_meta.update(sheet_meta)

        # Find where actual data headers start
        header_row_idx = _find_header_row(ws)
        if header_row_idx is None:
            warnings.append(f"Sheet '{sname}': could not detect header row")
            continue

        # Read headers
        raw_header_row = list(ws.iter_rows(
            min_row=header_row_idx,
            max_row=header_row_idx,
            values_only=True
        ))[0]

        raw_headers   = [str(h).strip() if h else "" for h in raw_header_row]
        norm_headers  = [h.lower().replace(" ", "_") for h in raw_headers]

        # Warn about unknown columns (extra columns are safe, just reported)
        if known_fields:
            for raw, norm in zip(raw_headers, norm_headers):
                if raw and norm not in known_fields:
                    warnings.append(f"Sheet '{sname}': unknown column '{raw}' will be ignored")

        # Determine which row to skip (type-hint row, required row — rows 2-3 of template)
        data_start_row = header_row_idx + 1
        # Skip optional annotation rows (they start with "Type:" or "Required")
        for check_row in range(header_row_idx + 1, header_row_idx + 4):
            row_vals = list(ws.iter_rows(min_row=check_row, max_row=check_row, values_only=True))[0]
            first_val = str(row_vals[0] or "").strip().lower() if row_vals else ""
            if first_val.startswith("type:") or first_val in ("required", "optional", "✓", "○"):
                data_start_row = check_row + 1
            else:
                break

        # Read data rows
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not any(row):
                continue   # Skip fully empty rows

            row_dict: dict[str, Any] = {}
            for norm_header, cell_val in zip(norm_headers, row):
                if norm_header:  # Skip columns with empty headers
                    row_dict[norm_header] = cell_val

            rows.append(row_dict)

        sheets[sname] = ParsedSheet(
            name        = sname,
            headers     = norm_headers,
            raw_headers = raw_headers,
            rows        = rows,
            meta        = sheet_meta,
        )

    return ParseResult(sheets=sheets, meta=file_meta, warnings=warnings)
