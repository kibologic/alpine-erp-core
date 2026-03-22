import io
from datetime import date
from typing import List, Optional

from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font
from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from core.db import get_session
from core.auth import get_current_tenant, verify_internal_token
from core.auth_deps import get_current_user_optional
from core.limits import LimitEnforcer, get_limit_enforcer
from core.models import Product, StockMovement, Supplier
from . import schemas, service

router = APIRouter(
    prefix="/inventory",
    tags=["Inventory"],
    dependencies=[Depends(verify_internal_token)],
)


@router.get("/categories", response_model=List[schemas.CategoryResponse])
async def list_categories(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    return await service.get_categories(session, tenant_id)


@router.post("/categories", response_model=schemas.CategoryResponse)
async def create_category(
    data: schemas.CategoryCreate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    return await service.create_category(session, tenant_id, data)


@router.patch("/categories/{category_id}", response_model=schemas.CategoryResponse)
async def update_category(
    category_id: str,
    data: schemas.CategoryUpdate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    from core.models import Category
    result = await session.execute(
        select(Category).where(Category.id == category_id, Category.tenant_id == tenant_id)
    )
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if data.name is not None:
        category.name = data.name
    await session.commit()
    await session.refresh(category)
    return category


@router.delete("/categories/{category_id}", status_code=204)
async def delete_category(
    category_id: str,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    from core.models import Category
    result = await session.execute(
        select(Category).where(Category.id == category_id, Category.tenant_id == tenant_id)
    )
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    await session.delete(category)
    await session.commit()
    return Response(status_code=204)


@router.get("/products", response_model=List[schemas.ProductResponse])
async def list_products(
    category_id: Optional[str] = None,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    return await service.get_products(session, tenant_id, category_id)


@router.post("/products", response_model=schemas.ProductResponse)
async def create_product(
    data: schemas.ProductCreate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
    limits: LimitEnforcer = Depends(get_limit_enforcer),
):
    await limits.check_product_limit()
    return await service.create_product(session, tenant_id, data)


@router.get("/products/{product_id}")
async def get_product(
    product_id: str,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        text("""
            SELECT
                p.id::text, p.tenant_id::text, p.category_id::text,
                p.name, p.sku, p.barcode, p.price, p.cost, p.tax_rate,
                p.is_active, p.created_at, p.extra_data,
                p.canonical_material_id, p.quality_tier,
                COALESCE(s.quantity, 0) AS stock_quantity
            FROM products p
            LEFT JOIN stock_on_hand s ON p.id = s.product_id AND p.tenant_id = s.tenant_id
            WHERE p.tenant_id = :t AND p.id::text = :id
        """),
        {"t": tenant_id, "id": product_id},
    )
    row = result.mappings().one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Product not found")
    return dict(row)


@router.patch("/products/{product_id}")
async def update_product(
    product_id: str,
    data: schemas.ProductUpdate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        select(Product).where(Product.id == product_id, Product.tenant_id == tenant_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(product, field, value)
    await session.commit()
    await session.refresh(product)
    # Re-fetch with stock_quantity
    stock_result = await session.execute(
        text("SELECT COALESCE(quantity, 0) FROM stock_on_hand WHERE product_id = :id AND tenant_id = :t"),
        {"id": product_id, "t": tenant_id},
    )
    stock_row = stock_result.one_or_none()
    stock_qty = float(stock_row[0]) if stock_row else 0.0
    return {
        "id": str(product.id), "tenant_id": str(product.tenant_id),
        "category_id": str(product.category_id) if product.category_id else None,
        "name": product.name, "sku": product.sku, "barcode": product.barcode,
        "price": float(product.price), "cost": float(product.cost),
        "tax_rate": float(product.tax_rate), "is_active": product.is_active,
        "created_at": product.created_at, "extra_data": product.extra_data,
        "canonical_material_id": product.canonical_material_id,
        "quality_tier": product.quality_tier, "stock_quantity": stock_qty,
    }


@router.delete("/products/{product_id}")
async def deactivate_product(
    product_id: str,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        select(Product).where(Product.id == product_id, Product.tenant_id == tenant_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    product.is_active = False
    await session.commit()
    return {"id": product_id, "deactivated": True}


@router.post("/adjust", response_model=schemas.StockMovementResponse)
async def adjust_stock(
    data: schemas.StockAdjustmentCreate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
    optional_user: dict | None = Depends(get_current_user_optional),
):
    created_by = optional_user["user_id"] if optional_user else None
    return await service.adjust_stock(session, tenant_id, data, user_id=created_by)


@router.get("/stock/{product_id}")
async def get_stock(
    product_id: str,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    qty = await service.get_stock_on_hand(session, tenant_id, product_id)
    return {"product_id": product_id, "quantity": qty}


@router.get("/valuation", response_model=schemas.InventoryValuationResponse)
async def get_valuation(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    return await service.get_inventory_valuation(session, tenant_id)


@router.get("/customers", response_model=List[schemas.CustomerResponse])
async def list_customers(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    return await service.get_customers(session, tenant_id)


@router.get("/movements")
async def list_movements(
    product_id: Optional[str] = None,
    reason: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    page_size = min(page_size, 200)
    offset = (page - 1) * page_size

    filters = "sm.tenant_id = :tenant_id"
    params: dict = {"tenant_id": tenant_id}

    if product_id:
        filters += " AND sm.product_id::text = :product_id"
        params["product_id"] = product_id
    if reason:
        filters += " AND sm.reason = :reason"
        params["reason"] = reason
    if date_from:
        filters += " AND sm.created_at >= :date_from"
        params["date_from"] = date_from
    if date_to:
        filters += " AND sm.created_at < :date_to::date + interval '1 day'"
        params["date_to"] = date_to

    count_result = await session.execute(
        text(f"SELECT COUNT(*) FROM stock_movements sm WHERE {filters}"),
        params,
    )
    total = count_result.scalar()

    items_result = await session.execute(
        text(f"""
            SELECT
                sm.id::text, sm.product_id::text,
                p.name AS product_name, p.sku AS product_sku,
                sm.quantity, sm.reason, sm.reference_type,
                sm.reference_id, sm.created_by::text, sm.created_at
            FROM stock_movements sm
            LEFT JOIN products p ON sm.product_id = p.id
            WHERE {filters}
            ORDER BY sm.created_at DESC
            LIMIT :limit OFFSET :offset
        """),
        {**params, "limit": page_size, "offset": offset},
    )
    rows = items_result.mappings().all()

    return {
        "items": [
            {
                "id": r["id"],
                "product_id": r["product_id"],
                "product_name": r["product_name"],
                "product_sku": r["product_sku"],
                "quantity": float(r["quantity"]),
                "reason": r["reason"],
                "reference_type": r["reference_type"],
                "reference_id": r["reference_id"],
                "created_by": r["created_by"],
                "created_at": r["created_at"].isoformat(),
            }
            for r in rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, -(-total // page_size)),  # ceiling division
    }


@router.get("/stock-levels")
async def list_stock_levels(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        text("""
            SELECT p.id, p.name, p.sku, p.price,
                   c.name as category,
                   COALESCE(s.quantity, 0) as stock_quantity
            FROM products p
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN stock_on_hand s ON p.id::text = s.product_id::text
                AND s.tenant_id::text = :tenant_id
            WHERE p.tenant_id::text = :tenant_id AND p.is_active = true
            ORDER BY p.name
        """),
        {"tenant_id": tenant_id}
    )
    rows = result.fetchall()
    return [
        {
            "id": str(r.id),
            "name": r.name,
            "sku": r.sku,
            "category": r.category,
            "price": float(r.price),
            "stock_quantity": float(r.stock_quantity),
            "status": "out_of_stock" if r.stock_quantity <= 0 else "low" if r.stock_quantity < 10 else "healthy",
        }
        for r in rows
    ]


def _supplier_dict(s: Supplier) -> dict:
    return {
        "id": str(s.id),
        "name": s.name,
        "email": s.email,
        "phone": s.phone,
        "address": s.address,
        "is_active": s.is_active,
        "created_at": s.created_at.isoformat(),
        "updated_at": s.updated_at.isoformat(),
    }


@router.get("/suppliers")
async def list_suppliers(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        select(Supplier)
        .where(Supplier.tenant_id == tenant_id, Supplier.is_active == True)
        .order_by(Supplier.name)
    )
    return [_supplier_dict(s) for s in result.scalars().all()]


class SupplierCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    is_active: bool = True


class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    is_active: Optional[bool] = None


@router.post("/suppliers")
async def create_supplier(
    data: SupplierCreate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    supplier = Supplier(
        tenant_id=tenant_id,
        name=data.name,
        email=data.email,
        phone=data.phone,
        address=data.address,
        is_active=data.is_active,
    )
    session.add(supplier)
    await session.commit()
    await session.refresh(supplier)
    return _supplier_dict(supplier)


@router.patch("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: str,
    data: SupplierUpdate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        select(Supplier).where(Supplier.id == supplier_id, Supplier.tenant_id == tenant_id)
    )
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(supplier, field, value)
    await session.commit()
    await session.refresh(supplier)
    return _supplier_dict(supplier)


@router.get("/stock-take/export")
async def export_stock_take(
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
):
    result = await session.execute(
        text("""
            SELECT p.sku, p.name, COALESCE(s.quantity, 0) as stock_on_hand
            FROM products p
            LEFT JOIN stock_on_hand s ON p.id::text = s.product_id::text
                AND s.tenant_id::text = :tenant_id
            WHERE p.tenant_id::text = :tenant_id AND p.is_active = true
            ORDER BY p.name
        """),
        {"tenant_id": tenant_id},
    )
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Take"

    headers = ["SKU", "Product Name", "Stock On Hand", "Counted Quantity", "Variance", "Notes"]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    for row in rows:
        ws.append([row.sku, row.name, float(row.stock_on_hand), None, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"stock-take-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/stock-take/import")
async def import_stock_take(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
    optional_user: dict | None = Depends(get_current_user_optional),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    if "Stock Take" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Sheet 'Stock Take' not found")

    ws = wb["Stock Take"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Pre-load all products for this tenant keyed by SKU
    result = await session.execute(
        select(Product).where(Product.tenant_id == tenant_id, Product.is_active == True)
    )
    products_by_sku = {p.sku: p for p in result.scalars().all()}

    processed = 0
    adjusted = 0
    skipped = []
    movement_ids = []

    for row in rows:
        if not any(row):
            continue

        sku, _name, stock_on_hand, counted_qty, _variance, notes = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )

        if counted_qty is None or counted_qty == "":
            continue

        processed += 1

        if sku not in products_by_sku:
            skipped.append(str(sku))
            continue

        stock_on_hand = float(stock_on_hand or 0)
        variance = float(counted_qty) - stock_on_hand

        if variance == 0:
            continue

        product = products_by_sku[sku]
        note_text = str(notes) if notes else "Stock take adjustment"

        movement = StockMovement(
            tenant_id=tenant_id,
            product_id=product.id,
            quantity=variance,
            reason="stock_take",
            reference_type=note_text,
            created_by=optional_user["user_id"] if optional_user else None,
        )
        session.add(movement)
        await session.flush()
        movement_ids.append(str(movement.id))
        adjusted += 1

    await session.commit()

    return {
        "processed": processed,
        "adjusted": adjusted,
        "skipped": skipped,
        "movements_created": movement_ids,
    }


@router.post("/customers", response_model=schemas.CustomerResponse)
async def create_customer(
    data: schemas.CustomerCreate,
    session: AsyncSession = Depends(get_session),
    tenant_id: str = Depends(get_current_tenant),
    limits: LimitEnforcer = Depends(get_limit_enforcer),
):
    await limits.check_customer_limit()
    return await service.create_customer(session, tenant_id, data)
